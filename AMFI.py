import openpyxl
import requests
import io
from openpyxl import load_workbook

excel = load_workbook('MFs.xlsx')
sheet_name = 'NAV'

idx = excel.sheetnames.index(sheet_name)
sheet = excel[sheet_name]
excel.remove(sheet)
excel.create_sheet(sheet_name, idx)
sheet = excel.active

url = 'https://www.amfiindia.com/spages/NAVAll.txt'
source = requests.get(url)
handle = io.StringIO(source.text)

for line in handle:
    words = line.split(';')
    if len(words) < 2: continue
    SchemeName = words[3]
    try:
        NAV = float(words[4])
    except:
        NAV = words[4]
    Date = words[5]
    sheet.append([SchemeName, NAV, Date])

excel.save('MFs.xlsx')